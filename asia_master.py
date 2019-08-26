import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from copy import copy
import os.path
from os import path
import time

start = time.time()

def as_text(value):
    if value is None:
        return ""
    return str(value)

def prCyan(skk): print("\033[96m {}\033[00m" .format(skk)) 
prCyan("\t\t Preparing Your File. Please Wait...")

asia = ['Armenia', 'Bangladesh', 'Cambodia', 'China', 'Hong Kong', 'India', 'Indonesia', 'Iran', 
'Israel', 'Japan', 'Jordan', 'Kuwait', 'Lebanon', 'Malaysia', 'Myanmar', 'Oman', 'Pakistan', 
'Philippines', 'Russia', 'Saudi Arabia', 'Singapore', 'South Korea', 'Sri Lanka', 'Taiwan', 
'Thailand', 'United Arab Emirates', 'Vietnam']

files = ['Armenia.xlsx', 'Bangladesh.xlsx', 'Cambodia.xlsx', 'China.xlsx', 'Hong Kong.xlsx',
'India.xlsx', 'Indonesia.xlsx', 'Iran.xlsx', 'Israel.xlsx', 'Japan.xlsx', 'Jordan.xlsx', 
'Kuwait.xlsx', 'Lebanon.xlsx', 'Malaysia.xlsx', 'Myanmar.xlsx', 'Oman.xlsx', 'Pakistan.xlsx', 
'Philippines.xlsx', 'Russia.xlsx', 'Saudi Arabia.xlsx', 'Singapore.xlsx', 'South Korea.xlsx', 
'Sri Lanka.xlsx', 'Taiwan.xlsx', 'Thailand.xlsx', 'United Arab Emirates.xlsx', 'Vietnam.xlsx']

master_sheet = "./master_asia.xlsx"

# Check if workbook already exists, create one if it does not exist
if not path.exists(master_sheet):

    # Create new workbook object
    wb = Workbook()

    # Create new excel sheets for each country
    for country in asia:
        # create country sheet if it does not already exist in the workbook
        if not country in wb.sheetnames:
            wb.create_sheet(country)
        # save workbook
        wb.save(master_sheet)

    # Remove Intial Blank Sheet if it exists 
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
        wb.save(master_sheet)


# load the new workbook where data will be copied
wb = load_workbook(master_sheet)

for country in asia:
    src_file = './' + country + '.xlsx'

    # original workbook is loaded in rawbook variable
    rawbook = load_workbook(src_file)
    rawsheet = rawbook.worksheets[0]
    
    wb[country].cell(row=1, column=1).value = "Serial No."
    wb[country].cell(row=1, column=2).value = "Location"
    wb[country].cell(row=1, column=3).value = "Regions"
    wb[country].cell(row=1, column=4).value = "Name of VC"
    wb[country].cell(row=1, column=5).value = "Year of Establishment"
    wb[country].cell(row=1, column=6).value = "Type of Investor"
    wb[country].cell(row=1, column=7).value = "Website"
    wb[country].cell(row=1, column=8).value = "LinkedIn"
    wb[country].cell(row=1, column=9).value = "Phone Details"
    wb[country].cell(row=1, column=10).value = "Contact Email"
    wb[country].cell(row=1, column=11).value = "Industry Vertical"
    wb[country].cell(row=1, column=12).value = "Investment Stage"
    wb[country].cell(row=1, column=13).value = "Number of Investments"
    wb[country].cell(row=1, column=14).value = "Number of Lead Investments"
    wb[country].cell(row=1, column=15).value = "Description"
    wb[country].cell(row=1, column=16).value = "Operating Status"
    wb[country].cell(row=1, column=17).value = "Company Type"
    wb[country].cell(row=1, column=18).value = "Number of Founders"
    wb[country].cell(row=1, column=19).value = "Number of Employees"
    wb[country].cell(row=1, column=20).value = "Total Funding Amount"
    wb[country].cell(row=1, column=21).value = "Number of Exits"
    wb[country].cell(row=1, column=22).value = "Last Funding Amount"


    row = 2 # keep track of row in output file
    rows = len(rawsheet['A'])

    i = 1 # keep track of rows for original file, as column is fixed at A or 1
    j = 1 # keep track of column for output file, as row will change after each loop

    max_len = 5

    while i <= rows+1:
        for j in range(1, 23):
            wb[country].cell(row=row, column=j).value = rawsheet.cell(row = i, column= 1).value
            wb[country].cell(row=row, column=j).hyperlink = rawsheet.cell(row = i, column= 1).hyperlink
            
            i += 1
        row += 1

    for j in range(1, 23):
        for i in range(1, int(rows/22 + 1)):

            cell_length = len(as_text(wb[country].cell(row=i, column=j).value))
            if cell_length > max_len:
                wb[country].column_dimensions[get_column_letter(j)].width = cell_length
                max_len = cell_length 
            else:
                wb[country].column_dimensions[get_column_letter(j)].width = max_len

    wb[country].delete_cols(1)

    wb.save(master_sheet)
    print(country+ " Done.")


end = time.time()

exec_time = end - start
print("\n\t\t Execution Time : ", exec_time, "\n")
